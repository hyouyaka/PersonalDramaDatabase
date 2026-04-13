from __future__ import annotations

import sqlite3
from pathlib import Path

from openpyxl import Workbook


ROOT = Path(__file__).resolve().parent
DB_PATH = ROOT / "DramasByCV.sqlite"
WORKBOOK_PATH = ROOT / "DramasByCV_merged.xlsx"
HEADERS = ["标题", "类型", "dramaids", "角色名", "总播放量", "平台"]


def build_workbook() -> None:
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    rows = conn.execute(
        """
        SELECT cv_name, title, genre, dramaids_text, role_names, total_play_count, platform
        FROM cv_works
        ORDER BY cv_name COLLATE NOCASE, platform, title COLLATE NOCASE, id
        """
    ).fetchall()
    conn.close()

    wb = Workbook()
    default_ws = wb.active
    default_ws.title = "Sheet1"

    sheet_map: dict[str, object] = {}
    for row in rows:
        cv_name = row["cv_name"]
        ws = sheet_map.get(cv_name)
        if ws is None:
            if default_ws.max_row == 1 and default_ws.max_column == 1 and default_ws["A1"].value is None and cv_name == rows[0]["cv_name"]:
                ws = default_ws
                ws.title = cv_name
            else:
                ws = wb.create_sheet(title=cv_name)
            ws.append(HEADERS)
            sheet_map[cv_name] = ws
        ws.append(
            [
                row["title"] or "",
                row["genre"] or "",
                row["dramaids_text"] or "",
                row["role_names"] or "",
                row["total_play_count"],
                row["platform"] or "",
            ]
        )

    if not rows:
        default_ws.append(HEADERS)

    for ws in wb.worksheets:
        if ws.max_row == 0:
            ws.append(HEADERS)
        for row_idx in range(2, ws.max_row + 1):
            ws.cell(row=row_idx, column=3).number_format = "@"

    wb.save(WORKBOOK_PATH)
    wb.close()


if __name__ == "__main__":
    build_workbook()
    print(f"Created: {WORKBOOK_PATH.name}")
